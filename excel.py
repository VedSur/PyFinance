from openpyxl import load_workbook, Workbook
from finance_management import Amount
from typing import NoReturn

def export_results_to_excel(file_path: str, report: dict[str, dict[str, Amount]]) -> NoReturn:
    workbook = Workbook()

    workbook.active['A1'] = 'Income'
    workbook.active['B1'] = 'Amount'

    workbook.active['D1'] = 'Expense'
    workbook.active['E1'] = 'Amount'

    workbook.active['G1'] = 'Asset'
    workbook.active['H1'] = 'Amount' 

    for i, item in enumerate(report['Income'].keys()):
        workbook.active[f'A{i+2}'] = item
        workbook.active[f'B{i+2}'] = report['Income'][item]

    for i, item in enumerate(report['Expense'].keys()):
        workbook.active[f'D{i+2}'] = item
        workbook.active[f'E{i+2}'] = report['Expense'][item]

    for i, item in enumerate(report['Investment'].keys()):
        workbook.active[f'G{i+2}'] = item
        workbook.active[f'H{i+2}'] = report['Investment'][item]
    
    workbook.save(file_path)

def export_transactions(file_path: str, transactions: list[tuple[str, str, Amount]]) -> NoReturn:
    workbook = Workbook()

    workbook.active['A1'] = 'Amount'
    workbook.active['B1'] = 'Name'
    workbook.active['C1'] = 'Type'

    for i, transaction in enumerate(transactions):
        workbook.active[f'A{i+2}'] = transaction[2]
        workbook.active[f'B{i+2}'] = str(transaction[0])
        workbook.active[f'C{i+2}'] = transaction[1]

    workbook.save(file_path)

def load_transactions(file_path: str) -> list[tuple[str, str, Amount]]:
    workbook = load_workbook(file_path)

    transactions: list[tuple[str, str, Amount]] = []

    i = 0
    while True:
        if workbook.active[f'A{i+2}'].value == None: break
        amount = Amount(workbook.active[f'A{i+2}'].value)
        name = str(workbook.active[f'B{i+2}'].value)
        transaction_type = workbook.active[f'C{i+2}'].value
        transactions.append((name, transaction_type, amount))
        i += 1
    return transactions